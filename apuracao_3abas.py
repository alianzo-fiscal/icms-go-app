# coding: utf-8
"""
apuracao_3abas.py
-----------------
Gera planilha com 3 abas:
  Aba 1: APURACAO ICMS  (por filial)
  Aba 2: BASE ENTRADAS  (consolidada por produto/CFOP/CST/aliq)
  Aba 3: BASE SAIDAS    (consolidada por produto/CFOP/CST/aliq)

Uso:
  python apuracao_3abas.py --entradas "Entradas*.xls" --saidas "Saidas*.xls Saidas*.csv"
                           --output "Apuracao.xlsx" --periodo "Junho/2026"
"""
import argparse, glob, sys, gc, subprocess, re, datetime
from pathlib import Path

import pandas as pd
import numpy as np
import xlrd
import openpyxl
from openpyxl import Workbook
from openpyxl.cell.cell import WriteOnlyCell
from openpyxl.styles import PatternFill, Font, Alignment, numbers, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constantes tributarias
# ---------------------------------------------------------------------------
ALIQ_INT_GO    = 19.0
ALIQ_INTER_NAC = 12.0
ALIQ_INTER_IMP = 4.0
ALIQ_DIFAL_SAI = 5.0   # 17% media destinos - 12% GO

CFOP_DIFAL_ENT = {'2551','2556','2910','2949','1551','1556'}
CFOP_SEM_CRED  = {'1908','1933','1551','2551'}
# Transferencias e devoluções: excluir do DIFAL saida EC87/2015
CFOP_EXCL_DIFAL_SAI = {'6151','6152','6153','6154','6155','6156',
                        '6201','6202','6208','6209','6210',
                        '6901','6902','6903','6949','6927'}
NCM_COSMET     = {'3303','3304','3305','3306','3307','3401','3402'}
NCM_COSMET_FULL= {'33030000','33041000','33042000','33043000','33049010',
                  '33049090','33051000','33052000','33053000','33059010',
                  '33059090','33061000','33062000','33069000','33071000',
                  '33072000','33073000','33074100','33074900','33079000',
                  '34011100','34011900','34012000','34013000',
                  '34021100','34021200','34021300','34021900','34022000','34029000'}

MESES_PT = {1:'Janeiro',2:'Fevereiro',3:'Marco',4:'Abril',5:'Maio',6:'Junho',
            7:'Julho',8:'Agosto',9:'Setembro',10:'Outubro',11:'Novembro',12:'Dezembro'}
MESES_NM = {'janeiro':1,'fevereiro':2,'marco':3,'abril':4,'maio':5,'junho':6,
            'julho':7,'agosto':8,'setembro':9,'outubro':10,'novembro':11,'dezembro':12}

# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------
def _fill(c): return PatternFill('solid', fgColor=c)
def _font(bold=False, color='000000', size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)
def _align(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

C_DARK = '1F3864'
C_MED  = '2E75B6'
C_HEAD = '1F4E79'
C_ALT  = 'F2F2F2'
C_YELL = 'FFEB9C'
C_GRN  = 'C6EFCE'
C_RED  = 'FFC7CE'
THIN   = Side(style='thin', color='BFBFBF')
BORD   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FMT_BRL = '#,##0.00'

# ---------------------------------------------------------------------------
# Leitura
# ---------------------------------------------------------------------------
def _num(v):
    if v is None or (isinstance(v, float) and np.isnan(v)): return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace(' ','').replace(',','.')
    try: return float(s)
    except: return 0.0

def _ler(caminho):
    ext = Path(caminho).suffix.lower()
    if ext == '.csv':
        return pd.read_csv(caminho, sep=';', encoding='latin-1', dtype=str, low_memory=False)
    try:
        return pd.read_excel(caminho, dtype=str, engine='xlrd')
    except:
        return pd.read_excel(caminho, dtype=str)

def carregar(padroes, tipo='saida'):
    frames = []
    for p in padroes:
        for f in sorted(glob.glob(p)):
            try:
                df = _ler(f)
                df.columns = [c.strip().upper() for c in df.columns]
                frames.append(df)
                print(f'  {tipo}: {Path(f).name} - {len(df):,} linhas', flush=True)
            except Exception as e:
                print(f'  ERRO {f}: {e}', flush=True)
    if not frames: return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)

def _periodo_arq(padroes):
    for p in padroes:
        for f in glob.glob(p):
            nm = Path(f).name.lower()
            for nome, num in MESES_NM.items():
                if nome in nm:
                    anos = re.findall(r'\b(20\d{2})\b', nm)
                    ano = int(anos[0]) if anos else datetime.datetime.now().year
                    return f'{MESES_PT[num]}/{ano}'
    return None

# ---------------------------------------------------------------------------
# Normalizacao de colunas
# ---------------------------------------------------------------------------
def _col(df, *nomes, default=0.0):
    for n in nomes:
        if n in df.columns: return df[n]
    return pd.Series([default]*len(df), index=df.index)

def norm_saidas(df):
    d = df.copy()
    d['_FILIAL']  = _col(d,'CODFILIAL','FILIAL').astype(str).str.strip()
    d['_CFOP']    = _col(d,'CODFISCAL','CFOP').astype(str).str.strip().str[:4]
    d['_CST']     = _col(d,'SITTRIBUT','CST','CSTICMS').astype(str).str.strip().str.zfill(2)
    d['_NCM']     = _col(d,'NCM').astype(str).str.strip().str.replace('.','',regex=False).str[:8]
    d['_PROD']    = _col(d,'CODPROD','CODIGO','CODITEM').astype(str).str.strip()
    d['_DESC']    = _col(d,'DESCRICAO','DESC','DESCR').astype(str).str.strip().str[:80]
    d['_ORIG']    = _col(d,'ORIGMERCTRIB','ORIGEM','ORIG').astype(str).str.strip().str[:1]
    d['_PERCICMS']= _col(d,'PERCICMS','ALIQICMS','ALIQUOTAICMS').apply(_num)
    d['_VLBASE']  = _col(d,'VLBASEICMS','BASEICMS','BASE').apply(_num)
    d['_VLICMS']  = _col(d,'VLICMS','ICMS').apply(_num)
    d['_VLITEM']  = _col(d,'VLITEM','VALOR','VL_ITEM').apply(_num)
    d['_VLCONT']  = _col(d,'VLCONTABIL','VL_CONTABIL','VLCONTAB').apply(_num)
    d['_QTCONT']  = _col(d,'QTCONT','QTD','QUANTIDADE').apply(_num)
    d['_NUMNOTA'] = _col(d,'NUMNOTA','NOTA','NUMNF').astype(str).str.strip()
    return d

def norm_entradas(df):
    d = df.copy()
    d['_FILIAL']  = _col(d,'CODFILIAL','FILIAL').astype(str).str.strip()
    d['_CFOP']    = _col(d,'CODFISCAL','CFOP').astype(str).str.strip().str[:4]
    d['_CST']     = _col(d,'CST','CSTICMS','SITTRIBUT').astype(str).str.strip().str.zfill(2)
    d['_NCM']     = _col(d,'NCM').astype(str).str.strip().str.replace('.','',regex=False).str[:8]
    d['_PROD']    = _col(d,'CODPROD','CODIGO','CODITEM').astype(str).str.strip()
    d['_DESC']    = _col(d,'DESCRICAO','DESC','DESCR').astype(str).str.strip().str[:80]
    d['_ORIG']    = _col(d,'ORIGMERCTRIB','ORIGEM','ORIG').astype(str).str.strip().str[:1]
    d['_PERCICMS']= _col(d,'PERCICMS','ALIQICMS').apply(_num)
    d['_VLBASE']  = _col(d,'VLBASEICMS','BASEICMS').apply(_num)
    d['_VLICMS']  = _col(d,'VLICMS','ICMS').apply(_num)
    d['_VLITEM']  = _col(d,'VLITEM','VALOR').apply(_num)
    d['_VLCONT']  = _col(d,'VLCONTABIL','VL_CONTABIL').apply(_num)
    d['_QTCONT']  = _col(d,'QTCONT','QTD','QUANTIDADE').apply(_num)
    d['_NUMNOTA'] = _col(d,'NUMNOTA','NOTA','NUMNF').astype(str).str.strip()
    return d

# ---------------------------------------------------------------------------
# Calculos de apuracao
# ---------------------------------------------------------------------------
def calc_debito(sai):
    mask = (sai['_CFOP'].str[:1].isin(['5','6'])) & (sai['_CST'].isin(['00','10','20','30']))
    grp      = sai[mask].groupby('_FILIAL')['_VLICMS'].sum()
    grp_base = sai[mask].groupby('_FILIAL')['_VLBASE'].sum()
    return {'total': float(grp.sum()), 'por_filial': grp,
            'base_total': float(grp_base.sum()), 'base_pf': grp_base}

def calc_credito(ent):
    mask = (ent['_CFOP'].str[:1].isin(['1','2'])) & \
           (ent['_CST'].isin(['00','20'])) & \
           (~ent['_CFOP'].isin(CFOP_SEM_CRED)) & \
           (ent['_VLICMS'] > 0)
    grp      = ent[mask].groupby('_FILIAL')['_VLICMS'].sum()
    grp_base = ent[mask].groupby('_FILIAL')['_VLBASE'].sum()
    return {'total': float(grp.sum()), 'por_filial': grp,
            'base_total': float(grp_base.sum()), 'base_pf': grp_base}

def calc_difal_ent(ent):
    mask = ent['_CFOP'].isin(CFOP_DIFAL_ENT) & ent['_CFOP'].str.startswith('2')
    grp_base = ent[mask].groupby('_FILIAL')['_VLBASE'].sum()
    grp_icms = ent[mask].groupby('_FILIAL')['_VLICMS'].sum()
    difal_f  = (grp_base * (ALIQ_INT_GO/100)) - grp_icms
    difal_f  = difal_f.clip(lower=0)
    return {'total': float(difal_f.sum()), 'por_filial': difal_f,
            'base_total': float(grp_base.sum()), 'base_pf': grp_base}

def calc_difal_sai(sai):
    mask    = (sai['_CFOP'].str.startswith('6') & sai['_CST'].isin(['00','20'])
             & (sai['_VLBASE'] > 0) & (~sai['_CFOP'].isin(CFOP_EXCL_DIFAL_SAI)))
    grp     = sai[mask].groupby('_FILIAL')['_VLBASE'].sum()
    difal_f = grp * (ALIQ_DIFAL_SAI/100)
    return {'total': float(difal_f.sum()), 'por_filial': difal_f,
            'base_total': float(grp.sum()), 'base_pf': grp}

def calc_protege(sai):
    # PROTEGE 15%: base reduzida com alíquota efetiva ≈ 11%
    # (redução de 19% para 11% ou de 12% para 11%) — usar alíquota efetiva, não só CST 20
    s = sai.copy()
    s['_ALIQ_EF'] = np.where(
        s['_VLCONT'] > 0,
        (s['_VLBASE'] / s['_VLCONT']) * s['_PERCICMS'],
        s['_PERCICMS']
    )
    m_p15 = (
        (s['_ALIQ_EF'] >= 10.0) & (s['_ALIQ_EF'] <= 11.9) &
        s['_PERCICMS'].isin([19.0, 12.0])
    )
    # Calcular diretamente por coluna — evita issue com groupby().apply() no pandas 2.x
    s_p15 = s[m_p15].copy()
    s_p15['_INTERVALO'] = (s_p15['_VLCONT'] - s_p15['_VLBASE']).clip(lower=0)
    p15_f = s_p15.groupby('_FILIAL')['_INTERVALO'].sum() * 0.15
    # PROTEGE 2%: todas as saídas com alíquota 21%
    m_p2 = s['_PERCICMS'] == 21.0
    p2_f = s[m_p2].groupby('_FILIAL')['_VLCONT'].sum() * 0.02
    del s, s_p15
    p15_tot = float(p15_f.sum()) if not p15_f.empty else 0.0
    p2_tot  = float(p2_f.sum())  if not p2_f.empty  else 0.0
    return {
        'total': p15_tot + p2_tot,
        'p15_por_filial': p15_f, 'p15_total': p15_tot,
        'p2_por_filial': p2_f, 'p2_total': p2_tot,
    }

# ---------------------------------------------------------------------------
# Consolidacao das bases
# ---------------------------------------------------------------------------
KEYS_SAI = ['_FILIAL','_PROD','_DESC','_NCM','_CFOP','_CST','_ORIG','_PERCICMS']
KEYS_ENT = ['_FILIAL','_PROD','_DESC','_NCM','_CFOP','_CST','_ORIG','_PERCICMS']

def base_saidas(sai):
    return sai.groupby(KEYS_SAI, dropna=False).agg(
        QTD=('_QTCONT','sum'), VL_ITEM=('_VLITEM','sum'),
        VL_CONTABIL=('_VLCONT','sum'), BASE_ICMS=('_VLBASE','sum'),
        ICMS=('_VLICMS','sum'), N_NOTAS=('_NUMNOTA','nunique')
    ).reset_index().rename(columns={
        '_FILIAL':'FILIAL','_PROD':'COD_PRODUTO','_DESC':'DESCRICAO',
        '_NCM':'NCM','_CFOP':'CFOP','_CST':'CST','_ORIG':'ORIGEM','_PERCICMS':'ALIQ_PCT'
    })

def base_entradas(ent):
    return ent.groupby(KEYS_ENT, dropna=False).agg(
        QTD=('_QTCONT','sum'), VL_ITEM=('_VLITEM','sum'),
        VL_CONTABIL=('_VLCONT','sum'), BASE_ICMS=('_VLBASE','sum'),
        ICMS=('_VLICMS','sum'), N_NOTAS=('_NUMNOTA','nunique')
    ).reset_index().rename(columns={
        '_FILIAL':'FILIAL','_PROD':'COD_PRODUTO','_DESC':'DESCRICAO',
        '_NCM':'NCM','_CFOP':'CFOP','_CST':'CST','_ORIG':'ORIGEM','_PERCICMS':'ALIQ_PCT'
    })


def _agg_sai(sai, mask):
    """Agrupa saídas filtradas pelo mask no padrão KEYS_SAI."""
    return sai[mask].groupby(KEYS_SAI, dropna=False).agg(
        QTD=('_QTCONT','sum'), VL_ITEM=('_VLITEM','sum'),
        VL_CONTABIL=('_VLCONT','sum'), BASE_ICMS=('_VLBASE','sum'),
        ICMS=('_VLICMS','sum'), N_NOTAS=('_NUMNOTA','nunique')
    ).reset_index().rename(columns={
        '_FILIAL':'FILIAL','_PROD':'COD_PRODUTO','_DESC':'DESCRICAO',
        '_NCM':'NCM','_CFOP':'CFOP','_CST':'CST','_ORIG':'ORIGEM','_PERCICMS':'ALIQ_PCT'
    })


def _base_difal_saida(sai):
    """Saídas interestaduais — base DIFAL EC 87/2015"""
    mask = (sai['_CFOP'].str.startswith('6') & sai['_CST'].isin(['00','20'])
           & (sai['_VLBASE'] > 0) & (~sai['_CFOP'].isin(CFOP_EXCL_DIFAL_SAI)))
    df = _agg_sai(sai, mask)
    df['DIFAL_SAIDA'] = (df['BASE_ICMS'] * ALIQ_DIFAL_SAI / 100).round(2)
    return df


def _base_protege15(sai):
    """Base PROTEGE 15% — alíq. efetiva ≈ 11% (nominal 19% ou 12%)"""
    s = sai.copy()
    s['_ALIQ_EF'] = np.where(
        s['_VLCONT'] > 0,
        (s['_VLBASE'] / s['_VLCONT']) * s['_PERCICMS'],
        s['_PERCICMS']
    )
    mask = (
        (s['_ALIQ_EF'] >= 10.0) & (s['_ALIQ_EF'] <= 11.9) &
        s['_PERCICMS'].isin([19.0, 12.0])
    )
    df = _agg_sai(s, mask)
    df['INTERVALO']     = (df['VL_CONTABIL'] - df['BASE_ICMS']).clip(lower=0).round(2)
    df['PROTEGE_15PCT'] = (df['INTERVALO'] * 0.15).round(2)
    del s
    return df


def _base_protege2(sai):
    """Base PROTEGE 2% — todas as saídas com alíquota 21%"""
    mask = sai['_PERCICMS'] == 21.0
    df = _agg_sai(sai, mask)
    df['PROTEGE_2PCT'] = (df['VL_CONTABIL'] * 0.02).round(2)
    return df


def _base_icms_comp(sai):
    """Base ICMS Complementar CFOP 5949 — 19% sobre VL_CONTABIL"""
    mask = sai['_CFOP'] == '5949'
    df = _agg_sai(sai, mask)
    df['ICMS_COMPLEMENTAR'] = (df['VL_CONTABIL'] * ALIQ_INT_GO / 100).round(2)
    return df


def _base_estorno(sai):
    """Base Estorno CFOP 5927 — 12% sobre BASE_ICMS"""
    mask = sai['_CFOP'] == '5927'
    df = _agg_sai(sai, mask)
    df['ICMS_ESTORNO'] = (df['BASE_ICMS'] * ALIQ_INTER_NAC / 100).round(2)
    return df


def _agg_ent(ent, mask):
    """Agrupa entradas filtradas pelo mask no padrao KEYS_ENT."""
    return ent[mask].groupby(KEYS_ENT, dropna=False).agg(
        QTD=('_QTCONT','sum'), VL_ITEM=('_VLITEM','sum'),
        VL_CONTABIL=('_VLCONT','sum'), BASE_ICMS=('_VLBASE','sum'),
        ICMS=('_VLICMS','sum'), N_NOTAS=('_NUMNOTA','nunique')
    ).reset_index().rename(columns={
        '_FILIAL':'FILIAL','_PROD':'COD_PRODUTO','_DESC':'DESCRICAO',
        '_NCM':'NCM','_CFOP':'CFOP','_CST':'CST','_ORIG':'ORIGEM','_PERCICMS':'ALIQ_PCT'
    })


def _base_difal_entrada(ent):
    """Entradas interestaduais sujeitas a DIFAL (uso/consumo e ativo imobilizado)."""
    mask = ent['_CFOP'].isin(CFOP_DIFAL_ENT) & ent['_CFOP'].str.startswith('2')
    df = _agg_ent(ent, mask)
    if df.empty:
        return df
    # Aliquota interestadual: 4% importado (ORIG==1), 12% nacional
    df['ALIQ_INTER'] = df['ORIGEM'].apply(lambda o: ALIQ_INTER_IMP if str(o).strip() == '1' else ALIQ_INTER_NAC)
    # DIFAL = max(0, BASE * 19% - ICMS_PAGO)
    df['DIFAL_ENTRADA'] = ((df['BASE_ICMS'] * ALIQ_INT_GO / 100) - df['ICMS']).clip(lower=0).round(2)
    return df


# ---------------------------------------------------------------------------
# Geracao da Aba 1 - APURACAO ICMS
# ---------------------------------------------------------------------------
def gerar_apuracao(output_path, periodo, filiais, debito, credito, difal_e, difal_s, protege):
    wb = Workbook()
    ws = wb.active
    ws.title = 'APURACAO ICMS'
    ws.sheet_properties.tabColor = C_DARK
    ws.freeze_panes = 'B4'

    fils = sorted(str(f) for f in filiais if f == f and f is not None)
    ncols = len(fils) + 2  # col A desc + filiais + total

    ws.column_dimensions['A'].width = 44
    for i in range(len(fils)):
        ws.column_dimensions[get_column_letter(i+2)].width = 18
    ws.column_dimensions[get_column_letter(ncols)].width = 18

    def titulo_bloco(row, txt, cor=C_DARK):
        ws.row_dimensions[row].height = 22
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row, 1, txt)
        c.font = _font(True, 'FFFFFF', 11)
        c.fill = _fill(cor)
        c.alignment = _align()

    def linha(row, desc, por_filial, total, bold=False, cor=None, fmt=FMT_BRL):
        ws.row_dimensions[row].height = 18
        c = ws.cell(row, 1, desc)
        c.font = _font(bold)
        c.alignment = _align('left')
        if cor: c.fill = _fill(cor)
        pf = por_filial if isinstance(por_filial, dict) else (por_filial.to_dict() if hasattr(por_filial,'to_dict') else {})
        for i, f in enumerate(fils):
            cc = ws.cell(row, i+2, round(pf.get(f, 0.0), 2))
            cc.number_format = fmt
            cc.alignment = _align('right')
            cc.font = _font(bold)
            if cor: cc.fill = _fill(cor)
        tc = ws.cell(row, ncols, round(float(total), 2))
        tc.number_format = fmt
        tc.alignment = _align('right')
        tc.font = _font(True)
        if cor: tc.fill = _fill(cor)

    # Linha 1 titulo
    ws.row_dimensions[1].height = 32
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws['A1']
    c.value = 'APURACAO DE ICMS  |  ' + periodo
    c.font = _font(True, 'FFFFFF', 14)
    c.fill = _fill(C_DARK)
    c.alignment = _align()

    # Linha 2 nota
    ws.row_dimensions[2].height = 15
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws['A2']
    c.value = 'Valores em R$  |  DIFAL saida: aliq media 17% destino  |  RCTE/GO Decreto 4.852/97'
    c.font = _font(False, '595959', 9, True)
    c.fill = _fill('F2F2F2')
    c.alignment = _align('left')

    # Linha 3 cabecalho colunas
    ws.row_dimensions[3].height = 22
    c = ws.cell(3, 1, 'DESCRICAO')
    c.font = _font(True, 'FFFFFF', 10)
    c.fill = _fill(C_HEAD)
    c.alignment = _align()
    for i, f in enumerate(fils):
        cc = ws.cell(3, i+2, 'FILIAL ' + str(f))
        cc.font = _font(True, 'FFFFFF', 10)
        cc.fill = _fill(C_HEAD)
        cc.alignment = _align()
    tc = ws.cell(3, ncols, 'TOTAL GRUPO')
    tc.font = _font(True, 'FFFFFF', 10)
    tc.fill = _fill(C_HEAD)
    tc.alignment = _align()

    row = 4
    # I - DEBITOS
    titulo_bloco(row, 'I.  DEBITOS DE ICMS - SAIDAS'); row += 1
    linha(row, 'Base de Calculo das Saidas', debito.get('base_pf',{}), debito.get('base_total',0)); row += 1
    linha(row, 'ICMS Debito (Saidas)', debito.get('por_filial',{}), debito.get('total',0), True); row += 1
    ws.row_dimensions[row].height = 6; row += 1

    # II - CREDITOS
    titulo_bloco(row, 'II.  CREDITOS DE ICMS - ENTRADAS'); row += 1
    linha(row, 'Base de Calculo das Entradas com Credito', credito.get('base_pf',{}), credito.get('base_total',0)); row += 1
    linha(row, 'ICMS Credito (Entradas)', credito.get('por_filial',{}), credito.get('total',0), True); row += 1
    ws.row_dimensions[row].height = 6; row += 1

    # III - SALDO PARCIAL
    saldo = debito.get('total',0) - credito.get('total',0)
    saldo_f = {}
    all_f = set(list(debito.get('por_filial',{}).keys()) + list(credito.get('por_filial',{}).keys()))
    for f in all_f:
        d = debito.get('por_filial',{})
        cr = credito.get('por_filial',{})
        dv = d.get(f,0) if isinstance(d,dict) else (d[f] if f in d.index else 0)
        cv = cr.get(f,0) if isinstance(cr,dict) else (cr[f] if f in cr.index else 0)
        saldo_f[f] = dv - cv
    linha(row, 'SALDO ANTES DO DIFAL  (Debito - Credito)', saldo_f, saldo, True, C_YELL); row += 1
    ws.row_dimensions[row].height = 6; row += 1

    # IV - DIFAL ENTRADA
    titulo_bloco(row, 'IV.  DIFAL ENTRADA (uso/consumo e ativo imobilizado)', C_MED); row += 1
    linha(row, 'Base DIFAL Entrada', difal_e.get('base_pf',{}), difal_e.get('base_total',0)); row += 1
    linha(row, 'DIFAL Entrada a Recolher (19% GO - 12%/4% inter.)', difal_e.get('por_filial',{}), difal_e.get('total',0), True); row += 1
    ws.row_dimensions[row].height = 6; row += 1

    # V - DIFAL SAIDA
    titulo_bloco(row, 'V.  DIFAL SAIDA EC 87/2015 (Consumidor Final Interestadual)', C_MED); row += 1
    linha(row, 'Base Vendas Interestaduais Consumidor Final', difal_s.get('base_pf',{}), difal_s.get('base_total',0)); row += 1
    linha(row, 'DIFAL Saida a Recolher (5% = 17% destino - 12% GO)', difal_s.get('por_filial',{}), difal_s.get('total',0), True); row += 1
    ws.row_dimensions[row].height = 6; row += 1

    # ICMS TOTAL A RECOLHER
    tot_icms = debito.get('total',0) - credito.get('total',0) + difal_e.get('total',0) + difal_s.get('total',0)
    tot_icms_f = {}
    for f in fils:
        def _g(d, f):
            if isinstance(d, dict): return d.get(f, 0.0)
            return float(d[f]) if f in d.index else 0.0
        tot_icms_f[f] = (_g(debito.get('por_filial',{}),f) - _g(credito.get('por_filial',{}),f)
                        + _g(difal_e.get('por_filial',{}),f) + _g(difal_s.get('por_filial',{}),f))
    cor_icms = C_RED if tot_icms > 0 else C_GRN
    linha(row, 'ICMS TOTAL A RECOLHER', tot_icms_f, tot_icms, True, cor_icms); row += 1
    ws.row_dimensions[row].height = 6; row += 1

    # VI - PROTEGE
    titulo_bloco(row, 'VI.  PROTEGE/GO - FUNDO SOCIAL (Lei 13.446/99)'); row += 1
    linha(row, 'PROTEGE 15% s/Intervalo (Aliq. Efetiva 11% — nominal 19% ou 12%)', protege.get('p15_por_filial',{}), protege.get('p15_total',0)); row += 1
    linha(row, 'PROTEGE 2% s/Valor — Todas as saidas com Aliquota 21%', protege.get('p2_por_filial',{}), protege.get('p2_total',0)); row += 1
    linha(row, 'TOTAL PROTEGE A RECOLHER',
          {f: protege.get('p15_por_filial',{}).get(f,0)+protege.get('p2_por_filial',{}).get(f,0) for f in fils},
          protege.get('total',0), True, C_YELL); row += 1
    ws.row_dimensions[row].height = 6; row += 1

    # VII - TOTAL GERAL
    tot_geral = tot_icms + protege.get('total',0)
    cor_geral = C_RED if tot_geral > 0 else C_GRN
    tot_geral_f = {f: tot_icms_f.get(f,0) + protege.get('p15_por_filial',{}).get(f,0) + protege.get('p2_por_filial',{}).get(f,0) for f in fils}
    linha(row, 'TOTAL GERAL  ICMS + PROTEGE/GO', tot_geral_f, tot_geral, True, cor_geral); row += 1

    wb.save(output_path)
    print('  Aba 1 APURACAO salva: ' + Path(output_path).name, flush=True)
    return tot_icms, tot_geral

# ---------------------------------------------------------------------------
# Geracao das Abas 2 e 3 com write_only (streaming)
# ---------------------------------------------------------------------------
COLS_BASE = ['FILIAL','COD_PRODUTO','DESCRICAO','NCM','CFOP','CST','ORIGEM','ALIQ_PCT',
             'QTD','VL_ITEM','VL_CONTABIL','BASE_ICMS','ICMS','N_NOTAS']
NUM_COLS_BASE = {'ALIQ_PCT','QTD','VL_ITEM','VL_CONTABIL','BASE_ICMS','ICMS','N_NOTAS'}
COL_WIDTHS = {'FILIAL':10,'COD_PRODUTO':14,'DESCRICAO':40,'NCM':12,'CFOP':8,
              'CST':7,'ORIGEM':8,'ALIQ_PCT':10,'QTD':12,'VL_ITEM':16,
              'VL_CONTABIL':16,'BASE_ICMS':16,'ICMS':16,'N_NOTAS':10}

NUM_COLS_EXTRA = {'DIFAL_SAIDA','DIFAL_ENTRADA','INTERVALO','PROTEGE_15PCT','PROTEGE_2PCT',
                  'ICMS_COMPLEMENTAR','ICMS_ESTORNO'}

def _header_row(ws, cols=None):
    if cols is None:
        cols = COLS_BASE
    cells = []
    for col in cols:
        c = WriteOnlyCell(ws, value=col)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = PatternFill('solid', fgColor=C_HEAD)
        c.alignment = Alignment(horizontal='center', vertical='center')
        cells.append(c)
    return cells

def _total_row(ws, df, cols=None):
    if cols is None:
        cols = COLS_BASE
    all_num = NUM_COLS_BASE | NUM_COLS_EXTRA
    cells = []
    for col in cols:
        if col in all_num:
            try:
                val = round(float(df[col].sum()), 2)
            except:
                val = 0.0
            c = WriteOnlyCell(ws, value=val)
            c.font = Font(bold=True)
            c.fill = PatternFill('solid', fgColor=C_YELL)
            c.number_format = FMT_BRL
        else:
            c = WriteOnlyCell(ws, value='TOTAL' if col == 'FILIAL' else '')
            c.font = Font(bold=True)
            c.fill = PatternFill('solid', fgColor=C_YELL)
        cells.append(c)
    return cells

def gerar_bases(output_path, periodo, base_ent, base_sai,
                base_difal=None, base_difal_ent=None, base_p15=None, base_p2=None,
                base_comp=None, base_estorno=None):
    wb = Workbook(write_only=True)

    def _escrever_aba(nome, titulo, df, tab_color, cols=None):
        if cols is None:
            cols = COLS_BASE
        ws = wb.create_sheet(nome)
        ws.sheet_properties.tabColor = tab_color
        for i, col in enumerate(cols):
            ws.column_dimensions[get_column_letter(i+1)].width = COL_WIDTHS.get(col, 16)
        # Titulo (linha 1)
        tc = WriteOnlyCell(ws, value=titulo)
        tc.font = Font(bold=True, color='FFFFFF', size=12)
        tc.fill = PatternFill('solid', fgColor=C_HEAD)
        tc.alignment = Alignment(horizontal='center', vertical='center')
        ws.append([tc])
        # Cabecalho (linha 2)
        ws.append(_header_row(ws, cols))
        # Dados
        n = len(df)
        df_out = df[[c for c in cols if c in df.columns]]
        for i, row in enumerate(df_out.itertuples(index=False)):
            ws.append(list(row))
            if i > 0 and i % 50000 == 0:
                print('    ' + nome + ': ' + str(i) + '/' + str(n) + ' linhas...', flush=True)
        # Total
        ws.append(_total_row(ws, df, cols))
        print('    ' + nome + ': ' + str(n) + ' linhas concluidas', flush=True)

    COLS_DIFAL     = COLS_BASE + ['DIFAL_SAIDA']
    COLS_DIFAL_ENT = COLS_BASE + ['ALIQ_INTER', 'DIFAL_ENTRADA']
    COLS_P15   = COLS_BASE + ['INTERVALO', 'PROTEGE_15PCT']
    COLS_P2    = COLS_BASE + ['PROTEGE_2PCT']
    COLS_COMP  = COLS_BASE + ['ICMS_COMPLEMENTAR']
    COLS_EST   = COLS_BASE + ['ICMS_ESTORNO']

    if len(base_ent) > 0:
        print('  Escrevendo BASE ENTRADAS (' + str(len(base_ent)) + ' linhas)...', flush=True)
        _escrever_aba('BASE ENTRADAS', 'BASE DE ENTRADAS CONSOLIDADA - ' + periodo, base_ent, '375623')
        del base_ent
        gc.collect()

    if len(base_sai) > 0:
        print('  Escrevendo BASE SAIDAS (' + str(len(base_sai)) + ' linhas)...', flush=True)
        _escrever_aba('BASE SAIDAS', 'BASE DE SAIDAS CONSOLIDADA - ' + periodo, base_sai, C_MED)
        del base_sai
        gc.collect()

    if base_difal is not None and len(base_difal) > 0:
        print('  Escrevendo BASE DIFAL (' + str(len(base_difal)) + ' linhas)...', flush=True)
        _escrever_aba('BASE DIFAL', 'BASE DIFAL SAIDA EC87/2015 - ' + periodo, base_difal, '7030A0', COLS_DIFAL)
        del base_difal; gc.collect()

    if base_difal_ent is not None and len(base_difal_ent) > 0:
        print('  Escrevendo BASE DIFAL ENTRADA (' + str(len(base_difal_ent)) + ' linhas)...', flush=True)
        _escrever_aba('BASE DIFAL ENTRADA', 'BASE DIFAL ENTRADA (DIFAL Compra) - ' + periodo, base_difal_ent, '4472C4', COLS_DIFAL_ENT)
        del base_difal_ent; gc.collect()

    if base_p15 is not None and len(base_p15) > 0:
        print('  Escrevendo BASE PROTEGE 15% (' + str(len(base_p15)) + ' linhas)...', flush=True)
        _escrever_aba('BASE PROTEGE 15%', 'BASE PROTEGE 15% - Aliq. Efetiva 11% - ' + periodo, base_p15, 'C00000', COLS_P15)
        del base_p15; gc.collect()

    if base_p2 is not None and len(base_p2) > 0:
        print('  Escrevendo BASE PROTEGE 2% (' + str(len(base_p2)) + ' linhas)...', flush=True)
        _escrever_aba('BASE PROTEGE 2%', 'BASE PROTEGE 2% - Saidas Aliq. 21% - ' + periodo, base_p2, 'FF7C00', COLS_P2)
        del base_p2; gc.collect()

    if base_comp is not None and len(base_comp) > 0:
        print('  Escrevendo BASE ICMS COMPL 5949 (' + str(len(base_comp)) + ' linhas)...', flush=True)
        _escrever_aba('BASE ICMS COMPL 5949', 'BASE ICMS COMPLEMENTAR CFOP 5949 (19%) - ' + periodo, base_comp, '00B050', COLS_COMP)
        del base_comp; gc.collect()

    if base_estorno is not None and len(base_estorno) > 0:
        print('  Escrevendo BASE ESTORNO 5927 (' + str(len(base_estorno)) + ' linhas)...', flush=True)
        _escrever_aba('BASE ESTORNO 5927', 'BASE ESTORNO CFOP 5927 (12%) - ' + periodo, base_estorno, '833C00', COLS_EST)
        del base_estorno; gc.collect()

    wb.save(output_path)
    print('  Abas bases salvas: ' + Path(output_path).name, flush=True)

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description='Apuracao ICMS - 3 abas')
    parser.add_argument('--saidas',   nargs='+', default=[])
    parser.add_argument('--entradas', nargs='+', default=[])
    parser.add_argument('--output',   default=None)
    parser.add_argument('--periodo',  default=None)
    args = parser.parse_args()

    pasta = Path(__file__).parent
    pads = args.saidas or [str(pasta/'saida*.xls'), str(pasta/'Saida*.xls'), str(pasta/'saida*.csv')]
    pade = args.entradas or [str(pasta/'entrada*.xls'), str(pasta/'Entrada*.xls')]

    print('=' * 60, flush=True)
    print('  Apuracao ICMS (3 abas)', flush=True)
    print('=' * 60, flush=True)

    print('\nCarregando saidas...', flush=True)
    sai_raw = carregar(pads, 'saida')
    if len(sai_raw) > 0:
        sai = norm_saidas(sai_raw)
        del sai_raw; gc.collect()
        print('  Total: ' + str(len(sai)) + ' registros | Filiais: ' + str(sai['_FILIAL'].nunique()), flush=True)
    else:
        sai = pd.DataFrame()

    print('\nCarregando entradas...', flush=True)
    ent_raw = carregar(pade, 'entrada')
    if len(ent_raw) > 0:
        ent = norm_entradas(ent_raw)
        del ent_raw; gc.collect()
        print('  Total: ' + str(len(ent)) + ' registros | Filiais: ' + str(ent['_FILIAL'].nunique()), flush=True)
    else:
        ent = pd.DataFrame()

    if len(sai) == 0 and len(ent) == 0:
        print('Nenhum arquivo encontrado.', flush=True)
        sys.exit(1)

    periodo = args.periodo
    if not periodo:
        periodo = _periodo_arq(pads + pade) or datetime.datetime.now().strftime('%B/%Y')
    print('\n  Periodo: ' + periodo, flush=True)

    filiais = set()
    if len(sai) > 0: filiais |= set(sai['_FILIAL'].unique())
    if len(ent) > 0: filiais |= set(ent['_FILIAL'].unique())
    fils = sorted(str(f) for f in filiais if f == f and f is not None)
    print('  Filiais: ' + ', '.join(fils), flush=True)

    print('\nCalculando apuracao...', flush=True)
    debito  = calc_debito(sai)   if len(sai) > 0 else {}
    credito = calc_credito(ent)  if len(ent) > 0 else {}
    difal_e = calc_difal_ent(ent) if len(ent) > 0 else {}
    difal_s = calc_difal_sai(sai) if len(sai) > 0 else {}
    protege = calc_protege(sai)  if len(sai) > 0 else {}

    print('Consolidando bases...', flush=True)
    base_sai_df  = base_saidas(sai)        if len(sai) > 0 else pd.DataFrame()
    base_ent_df  = base_entradas(ent)      if len(ent) > 0 else pd.DataFrame()
    base_difal_df = _base_difal_saida(sai) if len(sai) > 0 else pd.DataFrame()
    base_p15_df   = _base_protege15(sai)   if len(sai) > 0 else pd.DataFrame()
    base_p2_df    = _base_protege2(sai)    if len(sai) > 0 else pd.DataFrame()
    base_comp_df  = _base_icms_comp(sai)   if len(sai) > 0 else pd.DataFrame()
    base_difal_ent_df = _base_difal_entrada(ent) if len(ent) > 0 else pd.DataFrame()
    base_est_df   = _base_estorno(sai)     if len(sai) > 0 else pd.DataFrame()
    print('  BASE SAIDAS: ' + str(len(base_sai_df)) + ' linhas', flush=True)
    print('  BASE ENTRADAS: ' + str(len(base_ent_df)) + ' linhas', flush=True)
    print('  BASE DIFAL: ' + str(len(base_difal_df)) + ' linhas', flush=True)
    print('  BASE PROTEGE 15%: ' + str(len(base_p15_df)) + ' linhas', flush=True)
    print('  BASE PROTEGE 2%: ' + str(len(base_p2_df)) + ' linhas', flush=True)
    print('  BASE ICMS COMP 5949: ' + str(len(base_comp_df)) + ' linhas', flush=True)
    print('  BASE ESTORNO 5927: ' + str(len(base_est_df)) + ' linhas', flush=True)
    print('  BASE DIFAL ENTRADA: ' + str(len(base_difal_ent_df)) + ' linhas', flush=True)

    # Liberar DataFrames brutos
    del sai, ent
    gc.collect()

    periodo_nome = periodo.replace('/', ' ')
    output = args.output or str(pasta / ('Apuracao ICMS - ' + periodo_nome + '.xlsx'))

    # Arquivo temporario para apuracao (aba 1)
    tmp_apur = output.replace('.xlsx', '_tmp_apur.xlsx')
    tmp_base = output.replace('.xlsx', '_tmp_base.xlsx')

    print('\nGerando Aba 1 - APURACAO...', flush=True)
    tot_icms, tot_geral = gerar_apuracao(
        tmp_apur, periodo, fils, debito, credito, difal_e, difal_s, protege
    )

    print('\nGerando Abas de BASES...', flush=True)
    gerar_bases(tmp_base, periodo, base_ent_df, base_sai_df,
                base_difal=base_difal_df, base_difal_ent=base_difal_ent_df,
                base_p15=base_p15_df, base_p2=base_p2_df,
                base_comp=base_comp_df, base_estorno=base_est_df)
    del base_sai_df, base_ent_df, base_difal_df, base_difal_ent_df, base_p15_df, base_p2_df, base_comp_df, base_est_df
    gc.collect()

    # Combinar os dois arquivos
    combinar_script = Path(__file__).parent / 'apuracao-icms-go' / 'scripts' / 'combinar_xlsx.py'
    if combinar_script.exists():
        print('\nCombinando abas...', flush=True)
        r = subprocess.run(
            [sys.executable, str(combinar_script),
             '--principal', tmp_apur, '--adicional', tmp_base, '--output', output],
            capture_output=True, text=True
        )
        if r.returncode == 0:
            Path(tmp_apur).unlink(missing_ok=True)
            Path(tmp_base).unlink(missing_ok=True)
            print('  Arquivo final: ' + output, flush=True)
        else:
            print('  Erro ao combinar: ' + r.stderr[:300], flush=True)
            print('  Arquivos separados: ' + tmp_apur + ' e ' + tmp_base, flush=True)
    else:
        print('  combinar_xlsx.py nao encontrado. Arquivos: ' + tmp_apur + ' e ' + tmp_base, flush=True)

    print('\n' + '=' * 60, flush=True)
    brl = lambda v: 'R$ {:,.2f}'.format(v)
    print('  ICMS Debito (Saidas):       ' + brl(debito.get('total',0)), flush=True)
    print('  ICMS Credito (Entradas):    ' + brl(credito.get('total',0)), flush=True)
    dif = difal_e.get('total',0) + difal_s.get('total',0)
    print('  DIFAL Entrada + Saida:      ' + brl(dif), flush=True)
    print('  ICMS TOTAL A RECOLHER:      ' + brl(tot_icms), flush=True)
    print('  PROTEGE Total:              ' + brl(protege.get('total',0)), flush=True)
    print('  TOTAL GERAL ICMS+PROTEGE:   ' + brl(tot_geral), flush=True)
    print('=' * 60, flush=True)

if __name__ == '__main__':
    main()
